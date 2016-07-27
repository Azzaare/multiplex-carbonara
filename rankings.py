from tulip import *
from graphclass import *
from graphfuncs import *
from measures import *
import numpy as np
import scipy.stats as sc
import csv
import openpyxl as ox
import os.path
import random
import time                                                
#Hack to get common elements of two lists :
#keylist = list(set(l1) & set(l2))
    

################Ranking functions
#Function to rank nodes by their coeff.
def RankByCoeff(g, coeffname, limit = float('inf')):
    print coeffname
    coeff = g.graph[coeffname]
    H = {n : coeff[n] for n in g.graph.getNodes()} 
    l = sorted(H, key=H.__getitem__, reverse=True)

    #now remove all the elements of equal, bottom value
    botval = coeff[l[len(l)-1]]
    while coeff[l[len(l)-1]] == botval:
        l = l[0:len(l)-1]
    
    if limit != float('inf'):
        return l[:limit]
    else:
        return l


#Transforms a list of coeffs into their respective rank coeffs    
def CoeffToRank(g, names):
    for n in names:
        metric = g.graph[n]
        coeff = g.graph.getDoubleProperty("Rank"+n)
        
        l = RankByCoeff(g, n)
        
        coeff.setAllNodeValue(0)
    
        for i in range(len(l)):
            coeff[l[i]] = len(l)-i

#Calculates the difference in ranks between two measures and transforms it into a coeff
#how to judge order? Do all the orders possible
def CompareRank(g, names):
    CoeffToRank(g,names) #first convert to Rank
    

    #then compare
    for i in range(len(names)):
        for j in range(i):
            coeffij = g.graph.getDoubleProperty("RankDiff"+names[i]+"-"+names[j])
            coeffji = g.graph.getDoubleProperty("RankDiff"+names[j]+"-"+names[i])
            coeffi = g.graph["Rank"+names[i]]
            coeffj = g.graph["Rank"+names[j]]
            for n in g.graph.getNodes():
                coeffij[n] = coeffi[n]-coeffj[n]
                coeffji[n] = coeffj[n] - coeffi[n]
                

################COEFFICIENT CALCULATIONS
def KendallTauDistance(l1, l2):
    
    def round_to_one(f): #rounds to 1 the pathological cases due to floating point arithmetic errors
        if abs(f - 1.0) < 0.0000000001:
            return 1.0
        else:
            return f
    
    H1 = {}
    H2 = {}

    k=1
    for l in l1:
        H1[l] = k
        k+=1

    k=1
    for l in l2:
        H2[l] = k
        k+=1


    keylist = list(set(H1) & set(H2))
    #now create the two ranking lists
    ll1 = []
    ll2 = []
    
    for h in keylist:
        ll1+= [H1[h]]
        ll2+= [H2[h]]
        
    return round_to_one(sc.kendalltau(ll1,ll2)[0])


#PearsonCoeffs and SpearmanRankCoeffs give the same results
def PearsonCoeff(l1,l2): 
    #we need to first convert the list of nodes to a list of integers
    H1 = {}
    H2 = {}

    k=1
    for l in l1:
        H1[l] = k
        k+=1

    k=1
    for l in l2:
        H2[l] = k
        k+=1

    keylist = list(set(H1) & set(H2))
    #now create the two ranking lists
    ll1 = []
    ll2 = []
    
    for h in keylist:
        ll1+= [H1[h]]
        ll2+= [H2[h]]
        
    return np.corrcoef(ll1,ll2)[0,1]

def SpearmanRankCoeff(l1, l2):
    H1 = {}
    H2 = {}
    
    k=1
    for l in l1:
        H1[l] = k
        k+=1

    k=1
    for l in l2:
        H2[l] = k
        k+=1


    keylist = list(set(H1) & set(H2))

    #now create the two ranking lists
    ll1 = []
    ll2 = []
    
    for h in keylist:
        ll1+= [H1[h]]
        ll2+= [H2[h]]
        
    return sc.spearmanr(ll1,ll2)[0]

##########Functions to compare the measures

#CompareMeasures, by default using spearman corr coeff : 
def CompareMeasures(g, names, CorrelationFunc = SpearmanRankCoeff):
    l = [RankByCoeff(g,i) for i in names] #get all the lists

    for i in range(len(l)):
        for j in range(i):
            print names[i]+" x "+names[j],CorrelationFunc(l[i],l[j])

#Print top 20 coeffs            
def PrintTop20(g, names):
    l = [RankByCoeff(g,i) for i in names] 

    for i in range(len(l)):
        print names[i]
        for j in range(20):
            print j," : ", root.graph["node_property"][l[i][j]], " | ",g.graph[names[i]][l[i][j]]
    


def PrintDifferences(g, coeffname1, coeffname2, prec = "%.5f"):
    coeff1 = g.graph[coeffname1]
    coeff2 = g.graph[coeffname2]

    for n in g.graph.getNodes():
        if prec % coeff1[n] != prec % coeff2[n]:
        #if coeff1[n] != coeff2[n]:
            print n," : ",coeff1[n]," | ",coeff2[n]


#####Function to export the coeffs calculated into an excel file
#This function checks first if the excel file exists. If it does it doesn't recalculate whatever is already in it. Orders the coeff by alphabetical order

def ExportCoeffs(g, names, filename, CorrelationFunc = SpearmanRankCoeff):
    #we assume that all the coeffs included in the loaded excel file have already been calculated 
    def get_names(ws):
        i=2
        l=[]
        while not (ws.cell(row=i, column=1).value is None):
            l += [str(ws.cell(row=i, column=1).value)]
            i+=1
        return (i,l)
    
    #wb = ox.Workbook()
    alreadythere = []
    size = 2
    if os.path.isfile(filename+".xlsx") :
        wb = ox.load_workbook(filename+'.xlsx')
        ws = wb.active
        size,alreadythere = get_names(ws)
    else:
        wb = ox.Workbook()
        ws = wb.active

    ws.title = filename
    #get the new names to add
    toadd = list(set(names) - set(alreadythere))
    toadd = sorted(toadd, key=lambda x: (x[0].isdigit(), x[1].isdigit(), x))
    totalnames = list(set(alreadythere+names))
    print toadd

    #get all the rankings
    ranks = {n : RankByCoeff(g,n) for n in totalnames}
    
    #fill in the names of the measures to add and calculate the coeffs with the already added names
    for n in toadd:
        print "adding "+n
        size,alreadythere = get_names(ws)
        #add the name
        ws.cell(row = 1, column = size).value = ''.join([c for c in n if (c.isupper() or c.isdigit())])
        ws.cell(row = size, column = 1).value = n

        #add the name to alreadythere
        alreadythere+=[n]

        #calculate the coeff
        for i in range(len(alreadythere)):
            ws.cell(row = size, column = 2+i).value = CorrelationFunc(ranks[n],ranks[alreadythere[i]])

    wb.save(filename+".xlsx")


#This function doesn't check if an excel file exists, it recalculates everything and overwrites. Orders the coeffs in the order given by names.   
def ExportCoeffs_ordered(g, names, filename, CorrelationFunc = SpearmanRankCoeff):
    #get the workbook
    wb = ox.Workbook()
    ws = wb.active
    ws.title = filename
    
    #get all the rankings
    ranks = {n : RankByCoeff(g,n) for n in names}
    
    #fill in the names of the measures to add and calculate the coeffs with the already added names
    k = 2
    for n in names:
        print "adding "+n
        #add the name
        ws.cell(row = 1, column = k).value = ''.join([c for c in n if (c.isupper() or c.isdigit())])
        ws.cell(row = k, column = 1).value = n

        #calculate the coeff
        for i in range(k-1):
            ws.cell(row = k, column = 2+i).value = CorrelationFunc(ranks[n],ranks[names[i]])
        k+=1

    wb.save(filename+".xlsx")


#Function which reads in a file the coeffs and (assuming the coeffs in the file), gives the subgraph associated with this
def InducedCoeffs(names, oldfilename, newfilename):
    H = {}

    def get_names(ws):
        i=2
        l=[]
        while not (ws.cell(row=i, column=1).value is None):
            l += [str(ws.cell(row=i, column=1).value)]
            i+=1
        return (i,l)
    
    
    owb = ox.load_workbook(oldfilename+'.xlsx')
    ows = owb.active
    size,alreadythere = get_names(ows)

    print size
    print alreadythere

    for n in alreadythere :
        H[n] = {n : 0.0 for n in alreadythere} #Data structure which will hold all the values of the coeffs

    for i in range(2,size):
        for j in range(2,i+1):
            H[str(ows.cell(row=i, column=1).value)][str(ows.cell(row=j, column=1).value)] = str(ows.cell(row=i, column=j).value)
            H[str(ows.cell(row=j, column=1).value)][str(ows.cell(row=i, column=1).value)] = str(ows.cell(row=i, column=j).value)

    wb = ox.Workbook()
    ws = wb.active
    ws.title = newfilename

    k = 2
    for n in names:
        print "adding "+n
        #add the name
        ws.cell(row = 1, column = k).value = ''.join([c for c in n if (c.isupper() or c.isdigit())])
        ws.cell(row = k, column = 1).value = n

        #calculate the coeff
        for i in range(k-1):
            ws.cell(row = k, column = 2+i).value = H[n][names[i]]
        k+=1

    wb.save(newfilename+".xlsx")


#####Time comparison rankings
#A function to calculate the time taken to calculate a measure in a static way (but can be used with a dynamic measure). G must contain n subgraphs upon which the algo will be tried
def MeasureTime_static(g, measure, filename = ""):
    H = {}
    i=0
    for gk in g.subGraphs:
        print i
        graph = g[gk]
        ts = time.time()
        measure(graph)
        te = time.time()
        H[gk] = te - ts
        i+=1
        
    if filename != "": #if we decide to save it to an excel file
        wb = ox.Workbook()
        ws = wb.active
        ws.title = filename

        i = 1
        for gk in g.subGraphs:
            ws.cell(row = i, column = 1).value = g[gk].nNodes()
            ws.cell(row = i, column = 2).value = H[gk]
            i+=1

        wb.save(filename+".xlsx")

    return H


#TO TEST 
def MeasureTime_dynamic(g, measure_addnode, filename = "", threshold = 0):
    H = {}
    i=0
    for gk in g.subGraphs:
        print i
        graph = g[gk]

        gr = graph.addSubGraph(EmptyGraph)
        construction_order,edges = GetConstructionOrder(graph)

        totaltime = 0.0
        
        k=0
        for n in construction_order: #for all remaining nodes
            #add the node and its out edges
            gr.graph.addNode(n)
            for t in edges[n]:
                e = graph.graph.existEdge(n,t) #have to do this otherwise it creates duplicate edges...
                gr.graph.addEdge(e) 
            k+=1
            #update the flow of the graph while measuring the time it takes
            ts = time.time()
            measure_addnode(gr, n)
            te = time.time()

            totaltime+=(te-ts)
            
            if k%100 == 0:
                print k
        
        #now the graph's coeffs should be uptodate. Delete the subgraph, we don't need it anymore
        graph.delSubGraph(gr)
        H[gk] = totaltime
        i+=1
        print totaltime
        #if the last one was too long we stop here
        if totaltime > threshold:
            break
        
        
    if filename != "": #if we decide to save it to an excel file
        wb = ox.Workbook()
        ws = wb.active
        ws.title = filename

        i = 1
        for gk in H:
            ws.cell(row = i, column = 1).value = g[gk].nNodes()
            ws.cell(row = i, column = 2).value = H[gk]
            i+=1

        wb.save(filename+".xlsx")

    return H
    


#function which measures the time it takes to add one node and update the whole measure. Measure must be of type *_AddNode
def MeasureTime_addNode(g, measure_static, measure_addnode, filename = ""): 
    H = {}
    i=0

    node_order,edges = GetConstructionOrder(g) #get the construction order
    
    for gk in g.subGraphs:
        print i
        graph = g[gk]
                
        measure_static(graph)
        node_to_add = node_order[graph.nNodes()]
        
        #add the new node and its edges
        graph.graph.addNode(node_to_add)
        for t in edges[node_to_add]:
            e = g.graph.existEdge(node_to_add,t) #have to do this otherwise it creates duplicate edges...
            graph.graph.addEdge(e)

        #calculate the time it takes to update the measure
        #grraph = graph.addSubGraph(CloneGraph)
        ts = time.time()
        measure_addnode(graph,node_to_add)
        te = time.time()
        H[gk] = te - ts
        i+=1

        #graph.delSubGraph(grraph)
        
        graph.graph.delNode(node_to_add)
        
    if filename != "": #if we decide to save it to an excel file
        wb = ox.Workbook()
        ws = wb.active
        ws.title = filename

        i = 1
        for gk in g.subGraphs:
            
            ws.cell(row = i, column = 1).value = g[gk].nNodes()
            ws.cell(row = i, column = 2).value = H[gk]
            i+=1

        wb.save(filename+".xlsx")
        
    return H
    

